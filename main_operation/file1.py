from openpyxl import load_workbook


def alternative_crops(area, prev_crop):

    # data load
    wb = load_workbook('datas/data.xlsx')
    sheet_obj = wb.active

    data_dict = {}

    for i in range(2, sheet_obj.max_row + 1):
        key = sheet_obj.cell(row=i, column=1).value
        data_dict[key] = []

        for j in range(2, sheet_obj.max_column + 1):
            cell_obj = sheet_obj.cell(row=i, column=j)
            data_dict[key].append(cell_obj.value)

    # gross margin measurement other data table crops
    grossMargins = list()

    for i in data_dict:
        if prev_crop in i:
            continue

        _yield = data_dict[i][0]
        _market_price = data_dict[i][1]
        per_production_cost = data_dict[i][2]

        production = _yield * area
        gross_return = production * _market_price
        total_production_cost = area * per_production_cost
        gross_margin = round(gross_return - total_production_cost, 2)

        grossMargins.append((i, gross_margin))

    # gross margin for inputed crop
    if prev_crop not in data_dict.keys():
        prev_crop = "wheat"

    production = data_dict[prev_crop][0] * area
    gross_return = production * data_dict[prev_crop][1]
    total_production_cost = area * data_dict[prev_crop][2]
    prev_crop_gross_margin = round(gross_return - total_production_cost, 2)


    # measuring alternative crops
    expected_net_margin = []

    for i in grossMargins:
        ENMcc = i[1] - prev_crop_gross_margin
        if ENMcc > 0:
            expected_net_margin.append((i[0], ENMcc))

    alternative_crops = sorted(expected_net_margin, reverse=True, key=ENMc)

    return alternative_crops


def ENMc(ele):
    return ele[1]


# alternative_crops = alternative_crops_and_best_crop(are, prev_crop)
# print(alternative_crops)

# best_crop = file2.criteria_selection(pH, soil_type, temp, season)
# print(best_crop)

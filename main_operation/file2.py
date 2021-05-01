from openpyxl import load_workbook


# cross-check
def cross_check(data, pH, soil_type, temp, season):
    alternatives = []

    for i in data:
        count = 0

        if data[i][0][0] <= pH <= data[i][0][1]:
            count += 1

        if soil_type in data[i][1]:
            count += 1

        if data[i][2][0] <= temp <= data[i][2][1]:
            count += 1

        if season in data[i][3]:
            count += 1

        if count >= 3:
            alternatives.append((i, str(count*25)+"%"))

    return alternatives


def percent(ele):
    return ele[1]


# def criteria_selection(pH, soil_type, temp, season):
def best_crop(pH, soil_type, temp, season):

    data2 = load_workbook('datas/data2.xlsx')
    data2_obj = data2.active

    data2_dict = {}

    for i in range(2, data2_obj.max_row):
        key = data2_obj.cell(row=i, column=1).value
        data2_dict[key] = []

        for j in range(2, data2_obj.max_column+1):
            column = data2_obj.cell(row=i, column=j).value

            data2_dict[key].append(column.split(","))

    crops = cross_check(data2_dict, pH, soil_type, temp, season)
    suitable_crops = sorted(crops, key=percent)

    try:
        return suitable_crops[0]
    except:
        return None


# data = criteria_selection()


# print(data)
